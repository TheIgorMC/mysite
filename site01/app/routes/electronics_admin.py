"""
Electronics Admin Management Routes
Admin-only portal for electronics inventory, boards, BOMs, production jobs, and file management
"""
from flask import Blueprint, render_template, request, jsonify, current_app, redirect, url_for, flash
from flask_login import login_required, current_user
import requests
import csv
import io
import re
from functools import wraps
from app.api import OrionAPIClient

# Try to import openpyxl, provide helpful error if missing
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    # Note: Can't use current_app.logger here (module import time, no app context)
    import warnings
    warnings.warn("openpyxl not installed - Excel BOM parsing will not work", ImportWarning)

bp = Blueprint('electronics_admin', __name__, url_prefix='/admin/electronics')

# Also register API proxy routes under /electronics/api/ (like /archery/api/)
api_bp = Blueprint('electronics_api', __name__, url_prefix='/electronics/api')

def admin_required(f):
    """Decorator to ensure user is admin"""
    @wraps(f)
    @login_required
    def decorated_function(*args, **kwargs):
        if not current_user.is_admin:
            flash('Access denied. Admin privileges required.', 'error')
            return redirect(url_for('main.index'))
        return f(*args, **kwargs)
    return decorated_function

def api_request(endpoint, method='GET', data=None, params=None):
    """
    Make authenticated request to the API
    
    Args:
        endpoint: API endpoint path (e.g., '/api/elec/components')
        method: HTTP method
        data: Request body for POST/PATCH
        params: Query parameters
    
    Returns:
        Response JSON or None on error
    """
    api_base = current_app.config['API_BASE_URL']
    url = f"{api_base}{endpoint}"
    
    current_app.logger.info(f"[Electronics API] {method} {url}")
    if params:
        current_app.logger.info(f"[Electronics API] Params: {params}")
    if data:
        current_app.logger.info(f"[Electronics API] Data: {data}")
    
    headers = {
        'Content-Type': 'application/json',
        'Accept': 'application/json'
    }
    
    # Add Cloudflare Access headers if configured
    cf_id = current_app.config.get('CF_ACCESS_ID')
    cf_secret = current_app.config.get('CF_ACCESS_SECRET')
    if cf_id and cf_secret:
        headers['CF-Access-Client-Id'] = cf_id
        headers['CF-Access-Client-Secret'] = cf_secret
        current_app.logger.info(f"[Electronics API] Using CF Access authentication")
    else:
        current_app.logger.warning(f"[Electronics API] No CF Access credentials configured")
    
    try:
        if method == 'GET':
            response = requests.get(url, headers=headers, params=params, timeout=10)
        elif method == 'POST':
            response = requests.post(url, headers=headers, json=data, timeout=10)
        elif method == 'PATCH':
            response = requests.patch(url, headers=headers, json=data, timeout=10)
        elif method == 'DELETE':
            response = requests.delete(url, headers=headers, timeout=10)
        else:
            current_app.logger.error(f"[Electronics API] Invalid method: {method}")
            return None
        
        current_app.logger.info(f"[Electronics API] Response status: {response.status_code}")
        
        if response.status_code >= 400:
            current_app.logger.error(f"[Electronics API] Error response: {response.text}")
            # Forward error details (especially 422 validation errors) instead of swallowing them
            try:
                error_body = response.json()
            except Exception:
                error_body = {'detail': response.text}
            return {'_error': True, '_status': response.status_code, '_body': error_body}
        
        result = response.json()
        current_app.logger.info(f"[Electronics API] Success: {len(str(result))} bytes")
        return result
    except requests.exceptions.RequestException as e:
        current_app.logger.error(f"[Electronics API] Request failed: {str(e)}")
        return None

def api_result(result, error_msg='API request failed'):
    """Handle api_request result: return proper JSON response with correct status code.
    
    Handles three cases:
    - None: connection/timeout error → 500
    - Error dict (_error flag): API returned 4xx/5xx → forward status + body
    - Success: return JSON 200
    """
    if result is None:
        return jsonify({'error': error_msg}), 500
    if isinstance(result, dict) and result.get('_error'):
        return jsonify(result['_body']), result['_status']
    return jsonify(result)

@bp.route('/')
@admin_required
def index():
    """Electronics management portal main page"""
    storage_url = current_app.config.get('ELECTRONICS_STORAGE_URL', 'https://elec.orion-project.it')
    return render_template('admin/electronics.html', storage_url=storage_url)

# ============================================================================
# COMPONENTS API ENDPOINTS
# ============================================================================

@bp.route('/api/components', methods=['GET'])
@admin_required
def get_components():
    """Get components list with optional filters"""
    current_app.logger.info("[Electronics] get_components called")
    
    params = {
        'q': request.args.get('q', ''),
        'product_type': request.args.get('product_type', ''),
        'package': request.args.get('package', ''),
        'limit': request.args.get('limit', 100),
        'offset': request.args.get('offset', 0)
    }
    # Remove empty params
    params = {k: v for k, v in params.items() if v}
    
    current_app.logger.info(f"[Electronics] Calling API with params: {params}")
    result = api_request('/api/elec/components', params=params)
    return api_result(result, 'Failed to fetch components')

@bp.route('/api/components/search', methods=['GET'])
@admin_required
def search_components():
    """Smart component search (e.g., R0402 -> all 0402 resistors)"""
    q = request.args.get('q', '')
    result = api_request('/api/elec/components/search', params={'q': q})
    return api_result(result, 'Search failed')

@bp.route('/api/components', methods=['POST'])
@admin_required
def create_component():
    """Create new component"""
    data = request.get_json()
    result = api_request('/api/elec/components', method='POST', data=data)
    return api_result(result, 'Failed to create component')

@bp.route('/api/components/<component_id>', methods=['PATCH'])
@admin_required
def update_component(component_id):
    """Update component (e.g., quantity, price)"""
    data = request.get_json()
    result = api_request(f'/api/elec/components/{component_id}', method='PATCH', data=data)
    return api_result(result, 'Failed to update component')

@bp.route('/api/components/<component_id>', methods=['DELETE'])
@admin_required
def delete_component(component_id):
    """Delete component"""
    result = api_request(f'/api/elec/components/{component_id}', method='DELETE')
    return api_result(result, 'Failed to delete component')

# ============================================================================
# BOARDS API ENDPOINTS
# ============================================================================

@bp.route('/api/boards', methods=['GET'])
@admin_required
def get_boards():
    """Get boards list"""
    result = api_request('/api/elec/boards')
    return api_result(result, 'Failed to fetch boards')

@bp.route('/api/boards', methods=['POST'])
@admin_required
def create_board():
    """Create new board"""
    data = request.get_json()
    result = api_request('/api/elec/boards', method='POST', data=data)
    return api_result(result, 'Failed to create board')

@bp.route('/api/boards/<board_id>', methods=['PATCH'])
@admin_required
def update_board(board_id):
    """Update board info"""
    data = request.get_json()
    result = api_request(f'/api/elec/boards/{board_id}', method='PATCH', data=data)
    return api_result(result, 'Failed to update board')

@bp.route('/api/boards/<board_id>', methods=['DELETE'])
@admin_required
def delete_board(board_id):
    """Delete board"""
    result = api_request(f'/api/elec/boards/{board_id}', method='DELETE')
    return api_result(result, 'Failed to delete board')

@bp.route('/api/boards/<board_id>/bom', methods=['GET'])
@admin_required
def get_board_bom(board_id):
    """Get board's BOM"""
    result = api_request(f'/api/elec/boards/{board_id}/bom')
    return api_result(result, 'Failed to fetch BOM')

@bp.route('/api/boards/<board_id>/bom', methods=['POST'])
@admin_required
def update_board_bom(board_id):
    """Add/update components in board BOM"""
    data = request.get_json()
    result = api_request(f'/api/elec/boards/{board_id}/bom', method='POST', data=data)
    return api_result(result, 'Failed to update BOM')

@bp.route('/api/boards/<board_id>/bom/<component_id>', methods=['DELETE'])
@admin_required
def delete_bom_component(board_id, component_id):
    """Remove component from board BOM"""
    result = api_request(f'/api/elec/boards/{board_id}/bom/{component_id}', method='DELETE')
    return api_result(result, 'Failed to remove component from BOM')

@bp.route('/api/boards/<board_id>/upload_bom', methods=['POST'])
@admin_required
def upload_board_bom(board_id):
    """Upload BOM CSV - proxy to API"""
    # This endpoint expects the frontend to parse CSV and send component data
    # Format: [{"component_id": "...", "qty": 10}, ...]
    data = request.get_json()
    result = api_request(f'/api/elec/boards/{board_id}/upload_bom', method='POST', data=data)
    return api_result(result, 'Failed to upload BOM')

# ============================================================================
# JOBS API ENDPOINTS
# ============================================================================

@bp.route('/api/jobs', methods=['GET'])
@admin_required
def get_jobs():
    """Get production jobs list"""
    result = api_request('/api/elec/jobs')
    return api_result(result, 'Failed to fetch jobs')

@bp.route('/api/jobs', methods=['POST'])
@admin_required
def create_job():
    """Create new production job"""
    data = request.get_json()
    result = api_request('/api/elec/jobs', method='POST', data=data)
    return api_result(result, 'Failed to create job')

@bp.route('/api/jobs/<job_id>', methods=['GET'])
@admin_required
def get_job_details(job_id):
    """Get job details with BOM"""
    result = api_request(f'/api/elec/jobs/{job_id}')
    return api_result(result, 'Failed to fetch job details')

@bp.route('/api/jobs/<job_id>', methods=['PATCH'])
@admin_required
def update_job(job_id):
    """Update job status/quantity/due_date"""
    data = request.get_json()
    result = api_request(f'/api/elec/jobs/{job_id}', method='PATCH', data=data)
    return api_result(result, 'Failed to update job')

@bp.route('/api/jobs/<job_id>/check_stock', methods=['GET'])
@admin_required
def check_job_stock(job_id):
    """Check if sufficient stock for job"""
    result = api_request(f'/api/elec/jobs/{job_id}/check_stock')
    return api_result(result, 'Failed to check stock')

@bp.route('/api/jobs/<job_id>/reserve_stock', methods=['POST'])
@admin_required
def reserve_job_stock(job_id):
    """Reserve components for job (atomic operation)"""
    result = api_request(f'/api/elec/jobs/{job_id}/reserve_stock', method='POST')
    return api_result(result, 'Failed to reserve stock')

@bp.route('/api/jobs/<job_id>/missing_bom', methods=['GET'])
@admin_required
def get_missing_bom(job_id):
    """Get list of missing components for job"""
    result = api_request(f'/api/elec/jobs/{job_id}/missing_bom')
    return api_result(result, 'Failed to fetch missing components')

# ============================================================================
# FILES API ENDPOINTS
# ============================================================================

@bp.route('/api/boards/<board_id>/files', methods=['GET'])
@admin_required
def get_board_files(board_id):
    """Get list of files for a board"""
    result = api_request(f'/api/elec/boards/{board_id}/files')
    return api_result(result, 'Failed to fetch files')

@bp.route('/api/boards/<board_id>/files', methods=['POST'])
@admin_required
def register_board_file(board_id):
    """Register file metadata (file must already exist on nginx storage)"""
    data = request.get_json()
    result = api_request(f'/api/elec/boards/{board_id}/files', method='POST', data=data)
    return api_result(result, 'Failed to register file')

@bp.route('/api/boards/<board_id>/files/<file_id>', methods=['DELETE'])
@admin_required
def delete_board_file(board_id, file_id):
    """Delete file metadata"""
    result = api_request(f'/api/elec/boards/{board_id}/files/{file_id}', method='DELETE')
    return api_result(result, 'Failed to delete file')

@bp.route('/api/files/types', methods=['GET'])
@admin_required
def get_file_types():
    """Get supported file types"""
    result = api_request('/api/elec/files/types')
    return api_result(result, 'Failed to fetch file types')

# ============================================================================
# EXPORT ENDPOINTS
# ============================================================================

@bp.route('/api/bom/export', methods=['GET'])
@admin_required
def export_bom():
    """Export BOM as CSV"""
    params = {
        'format': request.args.get('format', 'csv'),
        'job_id': request.args.get('job_id', ''),
        'board_id': request.args.get('board_id', ''),
        'qty': request.args.get('qty', '')
    }
    # Remove empty params
    params = {k: v for k, v in params.items() if v}
    
    # For CSV export, we need to proxy the response
    api_base = current_app.config['API_BASE_URL']
    url = f"{api_base}/api/elec/bom/export"
    
    headers = {
        'Accept': 'text/csv'
    }
    
    cf_id = current_app.config.get('CF_ACCESS_ID')
    cf_secret = current_app.config.get('CF_ACCESS_SECRET')
    if cf_id and cf_secret:
        headers['CF-Access-Client-Id'] = cf_id
        headers['CF-Access-Client-Secret'] = cf_secret
    
    try:
        response = requests.get(url, headers=headers, params=params, timeout=30)
        response.raise_for_status()
        
        from flask import Response
        return Response(
            response.content,
            mimetype='text/csv',
            headers={
                'Content-Disposition': f'attachment; filename=bom_export.csv'
            }
        )
    except requests.exceptions.RequestException as e:
        current_app.logger.error(f"BOM export failed: {e}")
        return jsonify({'error': 'Failed to export BOM'}), 500

# ============================================================================
# PUBLIC API PROXY ROUTES (like /archery/api/)
# These routes proxy to the external API and are used by the frontend JavaScript
# ============================================================================

@api_bp.route('/components', methods=['GET'])
@login_required
def api_proxy_get_components():
    """Proxy: Get components list"""
    result = api_request('/api/elec/components', params=request.args.to_dict())
    return api_result(result, 'Failed to fetch components')

@api_bp.route('/components/search', methods=['GET'])
@login_required
def api_proxy_search_components():
    """Proxy: Smart component search"""
    params = request.args.to_dict()
    
    # API requires 'q' parameter - provide empty string if missing
    if 'q' not in params or not params['q']:
        params['q'] = ''
    
    result = api_request('/api/elec/components/search', params=params)
    return api_result(result, 'Search failed')

@api_bp.route('/components/types', methods=['GET'])
@login_required
def get_component_types():
    """Get unique component types from database"""
    try:
        api_client = OrionAPIClient()
        components = api_client.get_components(limit=10000)
        
        if not isinstance(components, list):
            return jsonify([])
        
        # Extract unique types/categories
        types = set()
        for comp in components:
            comp_type = comp.get('product_type') or comp.get('category')
            if comp_type:
                types.add(comp_type)
        
        return jsonify(sorted(list(types)))
    except Exception as e:
        current_app.logger.error(f"[Component Types] Error: {e}")
        return jsonify([])

@api_bp.route('/components/packages', methods=['GET'])
@login_required
def get_component_packages():
    """Get unique component packages/footprints from database"""
    try:
        api_client = OrionAPIClient()
        components = api_client.get_components(limit=10000)
        
        if not isinstance(components, list):
            return jsonify([])
        
        # Extract unique packages
        packages = set()
        for comp in components:
            pkg = comp.get('package')
            if pkg:
                packages.add(pkg)
        
        return jsonify(sorted(list(packages)))
    except Exception as e:
        current_app.logger.error(f"[Component Packages] Error: {e}")
        return jsonify([])

@api_bp.route('/components', methods=['POST'])
@login_required
def api_proxy_create_component():
    """Proxy: Create component"""
    result = api_request('/api/elec/components', method='POST', data=request.get_json())
    return api_result(result, 'Failed to create component')

@api_bp.route('/components/<component_id>', methods=['GET'])
@login_required
def api_proxy_get_component(component_id):
    """Proxy: Get single component by ID - fetches from full list"""
    # The Orion API doesn't have a GET endpoint for single components
    # So we fetch all and filter
    result = api_request('/api/elec/components', params={'limit': 10000})
    if result and isinstance(result, list):
        component = next((c for c in result if c.get('id') == int(component_id)), None)
        if component:
            return jsonify(component)
    return (jsonify({'error': 'Component not found'}), 404)

@api_bp.route('/components/<component_id>', methods=['PATCH'])
@login_required
def api_proxy_update_component(component_id):
    """Proxy: Update component"""
    result = api_request(f'/api/elec/components/{component_id}', method='PATCH', data=request.get_json())
    return api_result(result, 'Failed to update component')

@api_bp.route('/components/<component_id>', methods=['DELETE'])
@login_required
def api_proxy_delete_component(component_id):
    """Proxy: Delete component"""
    result = api_request(f'/api/elec/components/{component_id}', method='DELETE')
    return api_result(result, 'Failed to delete component')

@api_bp.route('/boards', methods=['GET'])
@login_required
def api_proxy_get_boards():
    """Proxy: Get boards list"""
    result = api_request('/api/elec/boards', params=request.args.to_dict())
    return api_result(result, 'Failed to fetch boards')

@api_bp.route('/boards', methods=['POST'])
@login_required
def api_proxy_create_board():
    """Proxy: Create board"""
    result = api_request('/api/elec/boards', method='POST', data=request.get_json())
    return api_result(result, 'Failed to create board')

@api_bp.route('/boards/<board_id>', methods=['GET'])
@login_required
def api_proxy_get_board(board_id):
    """Proxy: Get board details"""
    result = api_request(f'/api/elec/boards/{board_id}')
    return api_result(result, 'Failed to fetch board')

@api_bp.route('/boards/<board_id>/bom', methods=['GET'])
@login_required
def api_proxy_get_board_bom(board_id):
    """Proxy: Get board BOM"""
    result = api_request(f'/api/elec/boards/{board_id}/bom')
    return api_result(result, 'Failed to fetch BOM')

@api_bp.route('/boards/<board_id>/bom', methods=['POST'])
@login_required
def api_proxy_upload_bom(board_id):
    """Proxy: Upload/Save BOM"""
    result = api_request(f'/api/elec/boards/{board_id}/bom', method='POST', data=request.get_json())
    return api_result(result, 'Failed to save BOM')

@api_bp.route('/boards/<board_id>/bom/upload', methods=['POST'])
@login_required
def api_proxy_upload_bom_legacy(board_id):
    """Proxy: Upload BOM CSV (legacy endpoint)"""
    # Try the /upload endpoint first, fall back to regular /bom endpoint
    result = api_request(f'/api/elec/boards/{board_id}/bom/upload', method='POST', data=request.get_json())
    if not result:
        # Try without /upload suffix
        result = api_request(f'/api/elec/boards/{board_id}/bom', method='POST', data=request.get_json())
    return api_result(result, 'Failed to upload BOM')

@api_bp.route('/jobs', methods=['GET'])
@login_required
def api_proxy_get_jobs():
    """Proxy: Get jobs list"""
    try:
        result = api_request('/api/elec/jobs', params=request.args.to_dict())
        return api_result(result, 'Failed to fetch jobs - API returned no data')
    except Exception as e:
        current_app.logger.error(f"[Jobs Proxy] Exception: {e}", exc_info=True)
        return jsonify({'error': f'Failed to fetch jobs: {str(e)}'}), 500

@api_bp.route('/jobs', methods=['POST'])
@login_required
def api_proxy_create_job():
    """Proxy: Create job"""
    result = api_request('/api/elec/jobs', method='POST', data=request.get_json())
    return api_result(result, 'Failed to create job')

@api_bp.route('/jobs/<job_id>', methods=['GET'])
@login_required
def api_proxy_get_job(job_id):
    """Proxy: Get job details"""
    result = api_request(f'/api/elec/jobs/{job_id}')
    return api_result(result, 'Failed to fetch job')

@api_bp.route('/jobs/<job_id>/boards', methods=['POST'])
@login_required
def api_proxy_add_board_to_job(job_id):
    """Proxy: Add board to job"""
    result = api_request(f'/api/elec/jobs/{job_id}/boards', method='POST', data=request.get_json())
    return api_result(result, 'Failed to add board')

@api_bp.route('/jobs/<job_id>/check_stock', methods=['GET'])
@login_required
def api_proxy_check_job_stock(job_id):
    """Proxy: Check stock for job"""
    result = api_request(f'/api/elec/jobs/{job_id}/check_stock')
    return api_result(result, 'Failed to check stock')

@api_bp.route('/pnp', methods=['GET'])
@login_required
def api_proxy_get_pnp_files():
    """Proxy: Get PnP files list"""
    try:
        result = api_request('/api/elec/pnp', params=request.args.to_dict())
        # API might return None for 404 or empty array for no data
        if result is None:
            # Endpoint doesn't exist yet - return empty array
            current_app.logger.warning("[PnP Proxy] API endpoint not implemented yet")
            return jsonify([])
        return jsonify(result)
    except Exception as e:
        current_app.logger.error(f"[PnP Proxy] Exception: {e}", exc_info=True)
        return jsonify({'error': f'Failed to fetch PnP files: {str(e)}'}), 500

@api_bp.route('/pnp', methods=['POST'])
@login_required
def api_proxy_create_pnp():
    """Proxy: Upload PnP file"""
    result = api_request('/api/elec/pnp', method='POST', data=request.get_json())
    return api_result(result, 'Failed to upload PnP file')

@api_bp.route('/pnp/<pnp_id>', methods=['GET'])
@login_required
def api_proxy_get_pnp(pnp_id):
    """Proxy: Get PnP file details"""
    result = api_request(f'/api/elec/pnp/{pnp_id}')
    return api_result(result, 'Failed to fetch PnP file')

@api_bp.route('/pnp/<pnp_id>', methods=['DELETE'])
@login_required
def api_proxy_delete_pnp(pnp_id):
    """Proxy: Delete PnP file"""
    result = api_request(f'/api/elec/pnp/{pnp_id}', method='DELETE')
    return api_result(result, 'Failed to delete PnP file')

@api_bp.route('/jobs/<job_id>/reserve_stock', methods=['POST'])
@login_required
def api_proxy_reserve_stock(job_id):
    """Proxy: Reserve stock for job"""
    result = api_request(f'/api/elec/jobs/{job_id}/reserve_stock', method='POST')
    return api_result(result, 'Failed to reserve stock')

@api_bp.route('/jobs/<job_id>/missing_bom', methods=['GET'])
@login_required
def api_proxy_missing_bom(job_id):
    """Proxy: Get missing components BOM"""
    result = api_request(f'/api/elec/jobs/{job_id}/missing_bom')
    return api_result(result, 'Failed to generate BOM')

@api_bp.route('/files', methods=['GET'])
@login_required
def api_proxy_get_files():
    """Proxy: Get files list - aggregates files from all boards"""
    try:
        # Get board_id filter if provided
        board_id = request.args.get('board_id')
        
        if board_id:
            # Get files for specific board
            result = api_request(f'/api/elec/boards/{board_id}/files')
            # API might return None for empty or missing endpoint
            if result is None:
                return jsonify([])
        else:
            # Get all boards and aggregate their files
            boards = api_request('/api/elec/boards')
            if not boards:
                return jsonify([])
            
            all_files = []
            for board in boards:
                board_files = api_request(f'/api/elec/boards/{board["id"]}/files')
                # Handle empty or None responses
                if board_files:
                    # Add board info to each file
                    for file in board_files:
                        file['board_name'] = board.get('name', 'Unknown')
                        file['board_version'] = board.get('version', '')
                    all_files.extend(board_files)
            
            result = all_files
        
        return jsonify(result if result is not None else [])
    except Exception as e:
        current_app.logger.error(f"[Files Proxy] Exception: {e}", exc_info=True)
        return jsonify([])

@api_bp.route('/files/register', methods=['POST'])
@login_required
def api_proxy_register_file():
    """Proxy: Register file"""
    data = request.get_json()
    board_id = data.get('board_id')
    
    if not board_id:
        return jsonify({'error': 'board_id is required'}), 400
    
    # Remove board_id from data as it's in the URL
    data_copy = data.copy()
    data_copy.pop('board_id', None)
    
    result = api_request(f'/api/elec/boards/{board_id}/files', method='POST', data=data_copy)
    return api_result(result, 'Failed to register file')

@api_bp.route('/files/<file_id>', methods=['DELETE'])
@login_required
def api_proxy_delete_file(file_id):
    """Proxy: Delete file - requires board_id in query params"""
    board_id = request.args.get('board_id')
    
    if not board_id:
        return jsonify({'error': 'board_id query parameter is required'}), 400
    
    result = api_request(f'/api/elec/boards/{board_id}/files/{file_id}', method='DELETE')
    return api_result(result, 'Failed to delete file')

# ==================== ORDER IMPORTER ====================

@api_bp.route('/orders/parse', methods=['POST'])
@login_required
def parse_order_file():
    """Parse order file (LCSC CSV or Mouser XLS) and match components"""
    from werkzeug.utils import secure_filename
    import csv
    import io
    from datetime import datetime
    
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    supplier = request.form.get('supplier', '')
    order_date = request.form.get('order_date', '')
    
    # Validate inputs
    if not file or not supplier or not order_date:
        return jsonify({'error': 'Missing required fields'}), 400
    
    # Safely convert supplier to uppercase
    supplier = str(supplier).strip().upper() if supplier else ''
    if not supplier:
        return jsonify({'error': 'Supplier cannot be empty'}), 400
    
    try:
        # Read file content
        file_content = file.read()
        
        if not file_content:
            return jsonify({'error': 'File is empty'}), 400
        
        # Log file info for debugging
        current_app.logger.info(f"[Order Parse] File: {file.filename}, Size: {len(file_content)} bytes, Supplier: {supplier}")
        
        # Parse based on supplier
        if supplier == 'LCSC':
            items = parse_lcsc_csv(file_content)
        elif supplier == 'MOUSER':
            # Check file extension to determine format
            filename = file.filename.lower()
            if filename.endswith('.csv'):
                items = parse_mouser_csv(file_content)
            elif filename.endswith('.xlsx'):
                items = parse_mouser_xls(file_content)
            elif filename.endswith('.xls'):
                return jsonify({'error': 'Old Excel format (.xls) not supported. Please save as CSV or .xlsx format'}), 400
            else:
                return jsonify({'error': 'Unsupported file format. Mouser orders should be CSV or .xlsx'}), 400
        else:
            return jsonify({'error': 'Unsupported supplier'}), 400
        
        # Match components with database
        matched = []
        unmatched = []
        
        api_client = OrionAPIClient()
        components_result = api_client.get_components(limit=10000)
        all_components = components_result if isinstance(components_result, list) else []
        
        for item in items:
            # Try to match by seller_code or manufacturer_code
            matched_comp = None
            
            # Match by seller code first (safely)
            seller_code = item.get('seller_code', '').strip() if item.get('seller_code') else ''
            if seller_code:
                matched_comp = next((c for c in all_components 
                                   if str(c.get('seller_code', '')).strip().lower() == seller_code.lower()), None)
                
                # Also try 'supplier_code' field
                if not matched_comp:
                    matched_comp = next((c for c in all_components 
                                       if str(c.get('supplier_code', '')).strip().lower() == seller_code.lower()), None)
            
            # If not found, try manufacturer code (safely)
            if not matched_comp:
                mfg_code = item.get('manufacturer_code', '').strip() if item.get('manufacturer_code') else ''
                if mfg_code:
                    matched_comp = next((c for c in all_components 
                                       if str(c.get('manufacturer_code', '')).strip().lower() == mfg_code.lower()), None)
                    
                    # Also try 'mpn' field
                    if not matched_comp:
                        matched_comp = next((c for c in all_components 
                                           if str(c.get('mpn', '')).strip().lower() == mfg_code.lower()), None)
            
            if matched_comp:
                matched.append({
                    'component_id': matched_comp['id'],
                    'seller_code': item.get('seller_code'),
                    'manufacturer': item.get('manufacturer'),
                    'manufacturer_code': item.get('manufacturer_code'),
                    'package': item.get('package'),
                    'quantity': item['quantity'],
                    'unit_price': item['unit_price'],
                    'total_price': item['quantity'] * item['unit_price']
                })
            else:
                unmatched.append({
                    'seller_code': item.get('seller_code'),
                    'manufacturer': item.get('manufacturer'),
                    'manufacturer_code': item.get('manufacturer_code'),
                    'package': item.get('package'),
                    'description': item.get('description', ''),
                    'quantity': item['quantity'],
                    'unit_price': item['unit_price'],
                    'total_price': item['quantity'] * item['unit_price']
                })
        
        return jsonify({
            'supplier': supplier,
            'order_date': order_date,
            'matched': matched,
            'unmatched': unmatched
        })
        
    except Exception as e:
        current_app.logger.error(f"[Order Parse] Error: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500

@api_bp.route('/orders/import', methods=['POST'])
@login_required
def import_order():
    """Import order and update component stock quantities"""
    data = request.get_json()
    
    if not data or 'matched' not in data:
        return jsonify({'error': 'Invalid data'}), 400
    
    try:
        api_client = OrionAPIClient()
        
        # Get all components first to check current stock
        all_components_result = api_client.get_components(limit=10000)
        all_components = all_components_result if isinstance(all_components_result, list) else []
        components_by_id = {c['id']: c for c in all_components}
        
        updated_count = 0
        
        # Update stock for each matched component
        for item in data['matched']:
            component_id = item['component_id']
            quantity_to_add = item['quantity']
            unit_price = item['unit_price']
            
            # Get current component data
            current_comp = components_by_id.get(component_id)
            if not current_comp:
                current_app.logger.warning(f"[Order Import] Component {component_id} not found, skipping")
                continue
            
            # Calculate new stock (add to existing) - handle NULL/None values safely
            current_stock = current_comp.get('qty_left') or current_comp.get('stock_qty') or 0
            # Ensure current_stock is an integer (could be None, null, or string)
            try:
                current_stock = int(current_stock) if current_stock is not None else 0
            except (ValueError, TypeError):
                current_stock = 0
            
            new_stock = current_stock + quantity_to_add
            
            current_app.logger.info(f"[Order Import] Updating component {component_id}: {current_stock} + {quantity_to_add} = {new_stock}")
            
            # Update component stock and price
            update_data = {
                'qty_left': new_stock,
                'stock_qty': new_stock,  # Update both fields for compatibility
                'price': float(unit_price),
                'unit_price': float(unit_price)  # Update both fields for compatibility
            }
            
            api_client.update_component(component_id, **update_data)
            updated_count += 1
        
        return jsonify({'success': True, 'updated': updated_count})
        
    except Exception as e:
        current_app.logger.error(f"[Order Import] Error: {e}")
        return jsonify({'error': str(e)}), 500


def parse_lcsc_csv(file_content):
    """Parse LCSC CSV format"""
    items = []
    content = file_content.decode('utf-8')
    reader = csv.DictReader(io.StringIO(content))
    
    # USD to EUR conversion rate (approximate)
    USD_TO_EUR = 0.92  # Update this periodically or use an API
    
    for row in reader:
        # Try both Euro and Dollar currency symbols for Unit Price
        unit_price_eur = row.get('Unit Price(€)', '')
        unit_price_usd = row.get('Unit Price($)', '')
        
        unit_price = 0.0
        if unit_price_eur:
            # Price is already in euros
            unit_price = float(unit_price_eur.strip().replace(',', ''))
        elif unit_price_usd:
            # Price is in dollars, convert to euros
            unit_price_usd_float = float(unit_price_usd.strip().replace(',', ''))
            unit_price = unit_price_usd_float * USD_TO_EUR
        
        items.append({
            'seller_code': row.get('LCSC Part Number', '').strip(),
            'manufacturer_code': row.get('Manufacture Part Number', '').strip(),
            'manufacturer': row.get('Manufacturer', '').strip(),
            'package': row.get('Package', '').strip(),
            'description': row.get('Description', '').strip(),
            'quantity': int(row.get('Quantity', 0)),
            'unit_price': unit_price
        })
    
    return items


def parse_mouser_xls(file_content):
    """Parse Mouser XLSX format (Order History export)"""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel file parsing. Install it with: pip install openpyxl")
    
    items = []
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(file_content))
    except Exception as e:
        raise ValueError(f"Failed to read Excel file. Make sure it's a valid .xlsx file (not .xls or CSV). Error: {str(e)}")
    sheet = workbook.active
    
    # Find header row (usually row 1)
    headers = []
    header_row_idx = 1
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=5), start=1):
        row_values = [cell.value for cell in row]
        # Look for Mouser-specific headers
        if any(h and ('Mouser No' in str(h) or 'Mfr. No' in str(h)) for h in row_values):
            headers = [str(cell.value).strip() if cell.value else '' for cell in row]
            header_row_idx = row_idx
            current_app.logger.info(f"[Mouser XLSX] Found headers at row {row_idx}: {headers[:5]}...")
            break
    
    if not headers:
        raise ValueError("Could not find Mouser header row. Expected headers like 'Mouser No:', 'Mfr. No:', etc.")
    
    # Parse data rows
    for row in sheet.iter_rows(min_row=header_row_idx + 1):
        row_data = {headers[i]: cell.value for i, cell in enumerate(row) if i < len(headers)}
        
        # Try various field names for Mouser part number
        mouser_no = (row_data.get('Mouser No:') or 
                    row_data.get('Mouser No.') or 
                    row_data.get('Mouser Part No.') or 
                    row_data.get('Mouser No'))
        
        # Skip empty rows
        if not mouser_no:
            continue
        
        # Get manufacturer part number
        mfr_no = (row_data.get('Mfr. No:') or 
                 row_data.get('Mfr. No.') or 
                 row_data.get('Manufacturer Part No.') or 
                 row_data.get('Mfr No') or '')
        
        # Get description
        desc = (row_data.get('Desc.:') or 
               row_data.get('Desc.') or 
               row_data.get('Description') or '')
        
        # Get quantity
        qty_value = (row_data.get('Order Qty.') or 
                    row_data.get('Quantity') or 
                    row_data.get('Qty') or 0)
        try:
            quantity = int(qty_value) if qty_value else 0
        except (ValueError, TypeError):
            quantity = 0
        
        # Get price (EUR) - remove € symbol and convert comma to dot
        price_value = (row_data.get('Price (EUR)') or 
                      row_data.get('Unit Price') or 
                      row_data.get('Price') or '0')
        price_str = str(price_value).strip().replace('€', '').replace(',', '.').replace(' ', '')
        try:
            unit_price = float(price_str) if price_str else 0.0
        except (ValueError, TypeError):
            unit_price = 0.0
        
        # Get manufacturer name from dedicated field
        manufacturer = (row_data.get('Manufacturer:') or
                       row_data.get('Manufacturer') or
                       row_data.get('Mfr:') or
                       row_data.get('Mfr') or '')
        
        # Get product type/category from dedicated field if available
        product_type = (row_data.get('Type:') or
                       row_data.get('Type') or
                       row_data.get('Category:') or
                       row_data.get('Category') or '')
        
        items.append({
            'seller_code': str(mouser_no).strip(),
            'manufacturer_code': str(mfr_no).strip() if mfr_no else '',
            'manufacturer': str(manufacturer).strip() if manufacturer else '',
            'product_type': str(product_type).strip() if product_type else '',
            'package': '',  # Mouser order history doesn't include package
            'description': str(desc).strip() if desc else '',
            'quantity': quantity,
            'unit_price': unit_price
        })
    
    current_app.logger.info(f"[Mouser XLSX] Parsed {len(items)} items")
    return items


def parse_mouser_csv(file_content):
    """Parse Mouser CSV format (exported from Mouser packing list)"""
    items = []
    content = file_content.decode('utf-8')
    reader = csv.DictReader(io.StringIO(content))
    
    # USD to EUR conversion rate (approximate)
    USD_TO_EUR = 0.92
    
    # Log headers for debugging
    first_row = True
    
    for row in reader:
        if first_row:
            current_app.logger.info(f"[Mouser CSV] Headers: {list(row.keys())}")
            first_row = False
        
        # Try multiple field name variations for Mouser part number
        mouser_part = (row.get('Mouser Part No.') or 
                      row.get('Mouser Part No') or 
                      row.get('Mouser P/N') or 
                      row.get('Mouser No.') or 
                      row.get('Part Number') or '')
        
        # Skip empty rows
        if not mouser_part.strip():
            continue
        
        # Try to get manufacturer part number
        mfr_part = (row.get('Manufacturer Part No.') or 
                   row.get('Manufacturer Part No') or
                   row.get('Mfr. Part No.') or 
                   row.get('Mfr Part No') or
                   row.get('MPN') or '')
        
        # Try to get manufacturer name
        manufacturer = (row.get('Manufacturer') or 
                       row.get('Mfr.') or 
                       row.get('Mfr') or 
                       row.get('Brand') or '')
        
        # Try to get description
        description = (row.get('Description') or 
                      row.get('Product Description') or 
                      row.get('Desc') or '')
        
        # Try to get quantity (handle various formats)
        qty_str = (row.get('Quantity') or 
                  row.get('Qty') or 
                  row.get('Qty.') or 
                  row.get('Ordered Quantity') or '0')
        try:
            quantity = int(str(qty_str).strip().replace(',', ''))
        except (ValueError, TypeError):
            quantity = 0
        
        # Try to get price (Mouser uses different field names)
        unit_price_str = (row.get('Unit Price') or 
                         row.get('Price') or 
                         row.get('Unit Price ($)') or 
                         row.get('Unit Price (USD)') or
                         row.get('Price/Unit') or '0')
        unit_price_str = str(unit_price_str).strip().replace('$', '').replace('€', '').replace(',', '')
        
        # Convert USD to EUR
        try:
            unit_price_usd = float(unit_price_str) if unit_price_str else 0.0
            unit_price = unit_price_usd * USD_TO_EUR
        except (ValueError, TypeError):
            unit_price = 0.0
        
        items.append({
            'seller_code': str(mouser_part).strip(),
            'manufacturer_code': str(mfr_part).strip(),
            'manufacturer': str(manufacturer).strip(),
            'package': '',  # Mouser CSV usually doesn't include package
            'description': str(description).strip(),
            'quantity': quantity,
            'unit_price': unit_price
        })
    
    current_app.logger.info(f"[Mouser CSV] Parsed {len(items)} items")
    return items

# ===== STORAGE DIRECTORY LISTING =====

@api_bp.route('/storage/list', methods=['GET'])
@admin_required
def list_storage_directory():
    """
    Proxy endpoint to list files in a storage directory
    Bypasses CORS issues when scanning folders
    """
    folder_path = request.args.get('path', '')
    
    if not folder_path:
        return jsonify({'error': 'Path parameter required'}), 400
    
    # Get storage URL from config
    storage_url = current_app.config.get('ELECTRONICS_STORAGE_URL', 'https://elec.orion-project.it')
    
    # Clean up path - remove leading/trailing slashes
    folder_path = folder_path.strip('/')
    
    try:
        # Fetch directory listing
        url = f"{storage_url}/{folder_path}/"
        current_app.logger.info(f"[Storage List] Fetching: {url}")
        
        response = requests.get(url, timeout=10)
        
        if response.status_code != 200:
            current_app.logger.error(f"[Storage List] HTTP {response.status_code} from {url}")
            return jsonify({'error': f'Storage returned {response.status_code}'}), response.status_code
        
        html_content = response.text
        
        # Parse HTML to extract file links using regex (no external dependencies)
        import re
        
        # Match href attributes in <a> tags
        href_pattern = r'<a[^>]+href=["\']([^"\']+)["\']'
        matches = re.findall(href_pattern, html_content, re.IGNORECASE)
        
        files = []
        for href in matches:
            # Skip parent directory, query strings, and absolute paths
            if not href or href == '../' or href.startswith('?') or href.startswith('/'):
                continue
            
            # Remove trailing slash for directories
            filename = href.rstrip('/')
            
            # Skip subdirectories (containing /)
            if '/' in filename:
                continue
            
            # Skip empty or special entries
            if not filename or filename.startswith('.'):
                continue
            
            files.append(filename)
        
        current_app.logger.info(f"[Storage List] Found {len(files)} files in {folder_path}")
        return jsonify({'files': files, 'path': folder_path})
        
    except requests.RequestException as e:
        current_app.logger.error(f"[Storage List] Request failed: {e}")
        return jsonify({'error': f'Failed to fetch directory: {str(e)}'}), 500
    except Exception as e:
        current_app.logger.error(f"[Storage List] Error: {e}")
        return jsonify({'error': f'Error parsing directory: {str(e)}'}), 500

@api_bp.route('/storage/fetch', methods=['GET'])
@admin_required
def fetch_storage_file():
    """
    Proxy endpoint to fetch file contents from storage
    Bypasses CORS issues when loading BOM/PnP files
    """
    file_path = request.args.get('path', '')
    
    if not file_path:
        return jsonify({'error': 'Path parameter required'}), 400
    
    # Get storage URL from config
    storage_url = current_app.config.get('ELECTRONICS_STORAGE_URL', 'https://elec.orion-project.it')
    
    # Clean up path - remove leading/trailing slashes
    file_path = file_path.strip('/')
    
    try:
        # Fetch file content
        url = f"{storage_url}/{file_path}"
        current_app.logger.info(f"[Storage Fetch] Fetching: {url}")
        
        response = requests.get(url, timeout=30)
        
        if response.status_code != 200:
            current_app.logger.error(f"[Storage Fetch] HTTP {response.status_code} from {url}")
            return jsonify({'error': f'Storage returned {response.status_code}'}), response.status_code
        
        # Return file content as text (for CSV/text files)
        return response.text, 200, {'Content-Type': 'text/plain; charset=utf-8'}
        
    except requests.RequestException as e:
        current_app.logger.error(f"[Storage Fetch] Request failed: {e}")
        return jsonify({'error': f'Failed to fetch file: {str(e)}'}), 500
    except Exception as e:
        current_app.logger.error(f"[Storage Fetch] Error: {e}")
        return jsonify({'error': f'Error fetching file: {str(e)}'}), 500


@api_bp.route('/fetch-lcsc-price', methods=['GET'])
@login_required
def fetch_lcsc_price():
    """Fetch unit price from LCSC product page"""
    code = request.args.get('code', '').strip()
    if not code:
        return jsonify({'error': 'Missing code parameter'}), 400
    
    # Normalize: ensure it starts with C
    if not code.upper().startswith('C'):
        return jsonify({'error': 'Invalid LCSC code (must start with C)'}), 400
    
    url = f'https://www.lcsc.com/product-detail/{code}.html'
    current_app.logger.info(f'[LCSC Price] Fetching {url}')
    
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.9',
        }
        resp = requests.get(url, headers=headers, timeout=15)
        
        if resp.status_code != 200:
            current_app.logger.error(f'[LCSC Price] HTTP {resp.status_code}')
            return jsonify({'error': f'LCSC returned {resp.status_code}'}), 502
        
        html = resp.text
        
        # Extract prices from the price table
        # Pattern: look for euro prices like "€ 0.5194" or "€0.5194" in the price table cells
        price_pattern = re.compile(r'[€\u20ac]\s*([0-9]+\.?[0-9]*)\s*</span>')
        prices = price_pattern.findall(html)
        
        if not prices:
            # Try alternate pattern without euro sign (sometimes uses data attributes)
            price_pattern2 = re.compile(r'"unitPrice"\s*:\s*"?([0-9]+\.?[0-9]*)"?')
            prices = price_pattern2.findall(html)
        
        if not prices:
            # Try yet another pattern - look in JSON-LD or script data
            price_pattern3 = re.compile(r'price["\']?\s*[:=]\s*["\']?([0-9]+\.[0-9]{2,6})')
            prices = price_pattern3.findall(html)
        
        if prices:
            # Parse all price tiers
            price_values = [float(p) for p in prices if float(p) > 0]
            if price_values:
                # Return the first (1+ qty) unit price
                unit_price = price_values[0]
                current_app.logger.info(f'[LCSC Price] Found price for {code}: €{unit_price} ({len(price_values)} tiers)')
                return jsonify({
                    'code': code,
                    'unit_price': unit_price,
                    'price_tiers': price_values
                })
        
        current_app.logger.warning(f'[LCSC Price] No prices found in page for {code}')
        return jsonify({'error': 'Could not extract price from page'}), 404
        
    except requests.RequestException as e:
        current_app.logger.error(f'[LCSC Price] Request failed: {e}')
        return jsonify({'error': f'Failed to fetch: {str(e)}'}), 502
    except Exception as e:
        current_app.logger.error(f'[LCSC Price] Error: {e}')
        return jsonify({'error': f'Error: {str(e)}'}), 500

